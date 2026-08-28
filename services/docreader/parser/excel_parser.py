"""
Excel Parser Module

This module provides functionality to parse Excel files (.xlsx, .xls) into
structured Document objects with text content and chunks. It supports multiple
sheets and handles various Excel formats using pandas.
"""
import logging
import os
import re
import zipfile
from collections import defaultdict
from io import BytesIO
from typing import List
from xml.etree import ElementTree

import pandas as pd
from openpyxl.utils.cell import range_boundaries

from docreader.models.document import (
    Chunk,
    Document,
    SpreadsheetCell,
    SpreadsheetLocator,
    SpreadsheetRange,
    SpreadsheetTableIdentity,
    StructuredSourceUnit,
    StructuredSourceUnitKind,
)
from docreader.parser.base_parser import BaseParser
from docreader.parser.excel_convert import (
    convert_excel_to_xlsx_bytes,
    detect_excel_format,
    engine_for_format,
    normalize_excel_bytes,
)
from docreader.parser.xlsx_merge import fill_merged_cells_xlsx
from docreader.parser.xlsx_repair import repair_xlsx_bytes

logger = logging.getLogger(__name__)

# Pattern to detect Excel image function strings that should be excluded from
# parsed text content.  WPS uses =DISPIMG("ID",mode) to embed images in cells;
# when opened by other tools the formula may appear as plain text prefixed with
# "_xlfn." or "=".  Office 365 uses =_xlfn.IMAGE(url, ...) similarly.
# The _xlfn. prefix is optional — WPS may omit it (e.g. =DISPIMG("ID",1)).
_IMAGE_FUNC_RE = re.compile(
    r"^=?(_xlfn\.)?(DISPIMG|IMAGE)\(", re.IGNORECASE
)


def _positive_budget(name: str, default: int) -> int:
    try:
        value = int(os.environ.get(name, str(default)))
    except ValueError:
        return default
    return value if value > 0 else default


XLSX_MAX_MATERIALIZED_CELLS = _positive_budget("DOCREADER_XLSX_MAX_CELLS", 100_000)
XLSX_MAX_LOGICAL_ROW = _positive_budget("DOCREADER_XLSX_MAX_ROW", 100_000)
XLSX_MAX_LOGICAL_COLUMN = _positive_budget("DOCREADER_XLSX_MAX_COLUMN", 2_048)
XLSX_MAX_CELLS_PER_ROW = _positive_budget("DOCREADER_XLSX_MAX_CELLS_PER_ROW", 2_048)
XLSX_MAX_MERGES = _positive_budget("DOCREADER_XLSX_MAX_MERGES", 10_000)
XLSX_MAX_TABLES = _positive_budget("DOCREADER_XLSX_MAX_TABLES", 1_000)
XLSX_MAX_RANGE_CELLS = _positive_budget("DOCREADER_XLSX_MAX_RANGE_CELLS", 1_000_000)
XLSX_MAX_UNITS = _positive_budget("DOCREADER_XLSX_MAX_UNITS", 200_000)
XLSX_MAX_TEXT_BYTES = _positive_budget("DOCREADER_XLSX_MAX_TEXT_BYTES", 20 * 1024 * 1024)
XLSX_MAX_CELL_PAYLOAD_BYTES = _positive_budget(
    "DOCREADER_XLSX_MAX_CELL_PAYLOAD_BYTES", 64 * 1024 * 1024
)
XLSX_MAX_TABLE_CELL_SCANS = _positive_budget(
    "DOCREADER_XLSX_MAX_TABLE_CELL_SCANS", 2_000_000
)
XLSX_MAX_ARCHIVE_ENTRIES = _positive_budget("DOCREADER_XLSX_MAX_ARCHIVE_ENTRIES", 4_096)
XLSX_MAX_UNCOMPRESSED_BYTES = _positive_budget(
    "DOCREADER_XLSX_MAX_UNCOMPRESSED_BYTES", 100 * 1024 * 1024
)
XLSX_MAX_XML_PART_BYTES = _positive_budget(
    "DOCREADER_XLSX_MAX_XML_PART_BYTES", 50 * 1024 * 1024
)
XLSX_MAX_COMPRESSION_RATIO = _positive_budget(
    "DOCREADER_XLSX_MAX_COMPRESSION_RATIO", 100
)


class XlsxStructureLimitError(ValueError):
    """Stable fail-closed error for sparse or oversized XLSX structure."""


def _limit(name: str, actual: int, maximum: int) -> None:
    if actual > maximum:
        raise XlsxStructureLimitError(
            f"XLSX_STRUCTURE_LIMIT_EXCEEDED:{name}:{actual}>{maximum}"
        )


def _is_image_function(value: object) -> bool:
    """Return True if *value* looks like an embedded-image function string."""
    if not isinstance(value, str):
        return False
    return _IMAGE_FUNC_RE.match(value) is not None


class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls).

    This parser extracts text content from Excel files by processing all sheets
    and converting each row into a structured text format. Each row becomes a
    separate chunk with key-value pairs.

    Features:
        - Supports multiple sheets in a single Excel file
        - Automatically removes completely empty rows
        - Converts each row to "column: value" format
        - Creates individual chunks for each row for better granularity

    Example:
        >>> parser = ExcelParser()
        >>> with open("data.xlsx", "rb") as f:
        ...     content = f.read()
        ...     document = parser.parse_into_text(content)
        >>> print(document.content)
        Name: John,Age: 30,City: NYC
        Name: Jane,Age: 25,City: LA
    """

    def parse_into_text(self, content: bytes) -> Document:
        """Parse Excel file bytes into a Document object.

        Args:
            content: Raw bytes of the Excel file

        Returns:
            Document: Parsed document containing:
                - content: Full text with all rows from all sheets
                - chunks: List of Chunk objects, one per row

        Note:
            - Empty rows (all NaN values) are automatically skipped
            - Each row is formatted as: "col1: val1,col2: val2,..."
            - Chunks maintain sequential ordering across all sheets
        """
        chunks: List[Chunk] = []
        text: List[str] = []
        start, end = 0, 0

        # Tender XLSX follows the sparse structural path. Opening it through
        # pandas first can expand a single XFD1048576 cell into a huge rectangle.
        if detect_excel_format(content) == "xlsx":
            structured_units = _extract_xlsx_structured_units(content)
            for unit in structured_units:
                if unit.kind is not StructuredSourceUnitKind.TABLE_ROW or not unit.text:
                    continue
                content_row = unit.text + "\n"
                end += len(content_row)
                text.append(content_row)
                chunks.append(
                    Chunk(content=content_row, seq=len(chunks), start=start, end=end)
                )
                start = end
            return Document(
                content="".join(text),
                chunks=chunks,
                structured_source_units=structured_units,
            )

        excel_file = _open_excel_file(content, file_type=self.file_type)
        for excel_sheet_name in excel_file.sheet_names:
            df = _read_sheet_dataframe(excel_file, excel_sheet_name)
            df.dropna(how="all", inplace=True)
            for _, row in df.iterrows():
                page_content = [
                    f"{key}: {value}"
                    for key, value in row.items()
                    if pd.notna(value) and not _is_image_function(value)
                ]
                if not page_content:
                    continue
                content_row = ",".join(page_content) + "\n"
                end += len(content_row)
                text.append(content_row)
                chunks.append(
                    Chunk(content=content_row, seq=len(chunks), start=start, end=end)
                )
                start = end
        return Document(content="".join(text), chunks=chunks)


def _range_bounds(a1_range: str) -> tuple[int, int, int, int]:
    raw = range_boundaries(a1_range)
    values = [int(value) for value in raw if value is not None]
    if len(values) != 4:
        raise ValueError(f"XLSX range must be fully bounded: {a1_range}")
    return values[0], values[1], values[2], values[3]


def _range_model(a1_range: str) -> SpreadsheetRange:
    min_col, min_row, max_col, max_row = _range_bounds(a1_range)
    return SpreadsheetRange(
        a1_range=a1_range,
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
    )


def _cell_model(cell) -> SpreadsheetCell | None:
    value = cell.value
    if value is None or _is_image_function(value):
        return None
    return SpreadsheetCell(
        address=cell.coordinate,
        row=cell.row,
        column=cell.column,
        text=str(value),
    )


def _preflight_xlsx_range(a1_range: str, kind: str, max_area: int) -> None:
    try:
        min_col, min_row, max_col, max_row = _range_bounds(a1_range)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"XLSX_STRUCTURE_INVALID:{kind}_range") from exc
    _limit(f"{kind}_end_row", max_row, XLSX_MAX_LOGICAL_ROW)
    _limit(f"{kind}_end_column", max_col, XLSX_MAX_LOGICAL_COLUMN)
    _limit(f"{kind}_rows", max_row - min_row + 1, XLSX_MAX_LOGICAL_ROW)
    _limit(f"{kind}_columns", max_col - min_col + 1, XLSX_MAX_LOGICAL_COLUMN)
    _limit(
        f"{kind}_cells",
        (max_row - min_row + 1) * (max_col - min_col + 1),
        max_area,
    )


def _preflight_xlsx_structure(content: bytes) -> None:
    """Reject sparse/bomb geometry before openpyxl can materialize it."""
    try:
        archive = zipfile.ZipFile(BytesIO(content), "r")
    except (OSError, zipfile.BadZipFile) as exc:
        raise ValueError("XLSX_STRUCTURE_INVALID:container") from exc
    with archive:
        infos = archive.infolist()
        _limit("archive_entries", len(infos), XLSX_MAX_ARCHIVE_ENTRIES)
        total_uncompressed = 0
        worksheets = 0
        tables = 0
        cells = 0
        merges = 0
        cells_per_row: dict[int, int] = defaultdict(int)
        cell_references: set[tuple[str, str]] = set()
        archive_names: set[str] = set()
        for info in infos:
            name = info.filename.replace("\\", "/")
            if (
                name.startswith("/")
                or ".." in name.split("/")
                or name in archive_names
            ):
                raise ValueError("XLSX_STRUCTURE_INVALID:unsafe_path")
            archive_names.add(name)
            total_uncompressed += info.file_size
            _limit(
                "uncompressed_bytes",
                total_uncompressed,
                XLSX_MAX_UNCOMPRESSED_BYTES,
            )
            if info.file_size and (
                info.compress_size == 0
                or info.file_size > info.compress_size * XLSX_MAX_COMPRESSION_RATIO
            ):
                raise XlsxStructureLimitError(
                    "XLSX_STRUCTURE_LIMIT_EXCEEDED:compression_ratio"
                )
            is_worksheet = name.startswith("xl/worksheets/") and name.endswith(".xml")
            is_table = name.startswith("xl/tables/") and name.endswith(".xml")
            if not is_worksheet and not is_table:
                continue
            _limit("xml_part_bytes", info.file_size, XLSX_MAX_XML_PART_BYTES)
            if is_worksheet:
                worksheets += 1
            else:
                tables += 1
                _limit("tables", tables, XLSX_MAX_TABLES)
            try:
                with archive.open(info, "r") as stream:
                    for _, element in ElementTree.iterparse(stream, events=("end",)):
                        tag = element.tag.rsplit("}", 1)[-1]
                        if is_worksheet and tag == "dimension":
                            reference = element.attrib.get("ref")
                            if not reference:
                                raise ValueError("XLSX_STRUCTURE_INVALID:dimension")
                            _preflight_xlsx_range(
                                reference,
                                "logical_dimension",
                                10_000_000,
                            )
                        elif is_worksheet and tag == "c":
                            reference = element.attrib.get("r")
                            if not reference:
                                raise ValueError("XLSX_STRUCTURE_INVALID:cell")
                            try:
                                column, row, end_column, end_row = _range_bounds(reference)
                            except (TypeError, ValueError) as exc:
                                raise ValueError("XLSX_STRUCTURE_INVALID:cell") from exc
                            if (column, row) != (end_column, end_row) or (
                                name,
                                reference.upper(),
                            ) in cell_references:
                                raise ValueError("XLSX_STRUCTURE_INVALID:duplicate_cell")
                            cell_references.add((name, reference.upper()))
                            _limit("cell_row", row, XLSX_MAX_LOGICAL_ROW)
                            _limit("cell_column", column, XLSX_MAX_LOGICAL_COLUMN)
                            cells += 1
                            _limit(
                                "materialized_cells",
                                cells,
                                XLSX_MAX_MATERIALIZED_CELLS,
                            )
                            cells_per_row[row] += 1
                            _limit(
                                "cells_per_row",
                                cells_per_row[row],
                                XLSX_MAX_CELLS_PER_ROW,
                            )
                        elif is_worksheet and tag == "mergeCell":
                            reference = element.attrib.get("ref")
                            if not reference:
                                raise ValueError("XLSX_STRUCTURE_INVALID:merge")
                            merges += 1
                            _limit("merges", merges, XLSX_MAX_MERGES)
                            _preflight_xlsx_range(
                                reference,
                                "merge",
                                XLSX_MAX_RANGE_CELLS,
                            )
                        elif is_table and tag == "table":
                            reference = element.attrib.get("ref")
                            if not reference:
                                raise ValueError("XLSX_STRUCTURE_INVALID:table")
                            _preflight_xlsx_range(
                                reference,
                                "table",
                                XLSX_MAX_RANGE_CELLS,
                            )
                        element.clear()
            except ElementTree.ParseError as exc:
                raise ValueError("XLSX_STRUCTURE_INVALID:xml") from exc
        if worksheets == 0:
            raise ValueError("XLSX_STRUCTURE_INVALID:no_worksheet")


def _extract_xlsx_structured_units(content: bytes) -> List[StructuredSourceUnit]:
    """Return bounded deterministic units without rectangular expansion."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    _preflight_xlsx_structure(content)
    repaired = repair_xlsx_bytes(content)
    source = repaired if repaired is not None else content
    workbook = load_workbook(BytesIO(source), data_only=False, read_only=False)
    units: List[StructuredSourceUnit] = []
    total_text_bytes = 0
    total_cell_payload_bytes = 0

    def range_area(region: SpreadsheetRange, kind: str) -> None:
        rows = region.end_row - region.start_row + 1
        columns = region.end_column - region.start_column + 1
        _limit(f"{kind}_rows", rows, XLSX_MAX_LOGICAL_ROW)
        _limit(f"{kind}_columns", columns, XLSX_MAX_LOGICAL_COLUMN)
        _limit(f"{kind}_cells", rows * columns, XLSX_MAX_RANGE_CELLS)

    def contains(outer: SpreadsheetRange, inner: SpreadsheetRange) -> bool:
        return (
            outer.start_row <= inner.start_row
            and inner.end_row <= outer.end_row
            and outer.start_column <= inner.start_column
            and inner.end_column <= outer.end_column
        )

    def emit(unit: StructuredSourceUnit) -> None:
        nonlocal total_text_bytes, total_cell_payload_bytes
        _limit("units", len(units) + 1, XLSX_MAX_UNITS)
        total_text_bytes += len(unit.text.encode("utf-8"))
        _limit("text_bytes", total_text_bytes, XLSX_MAX_TEXT_BYTES)
        locator = unit.locator
        if isinstance(locator, SpreadsheetLocator):
            total_cell_payload_bytes += sum(
                len(cell.address.encode("utf-8")) + len(cell.text.encode("utf-8"))
                for cell in locator.cells
            )
            _limit(
                "cell_payload_bytes",
                total_cell_payload_bytes,
                XLSX_MAX_CELL_PAYLOAD_BYTES,
            )
        units.append(unit)

    try:
        for sheet_ordinal, sheet in enumerate(workbook.worksheets):
            max_row = max(1, int(sheet.max_row))
            max_column = max(1, int(sheet.max_column))
            _limit("logical_row", max_row, XLSX_MAX_LOGICAL_ROW)
            _limit("logical_column", max_column, XLSX_MAX_LOGICAL_COLUMN)

            raw_cells = getattr(sheet, "_cells", {})
            if not isinstance(raw_cells, dict):
                raise ValueError("XLSX_STRUCTURE_INVALID:materialized_cells")
            _limit(
                "materialized_cells",
                len(raw_cells),
                XLSX_MAX_MATERIALIZED_CELLS,
            )
            all_cells = sorted(
                (
                    model
                    for cell in raw_cells.values()
                    if (model := _cell_model(cell)) is not None
                ),
                key=lambda cell: (cell.row, cell.column, cell.address),
            )
            rows: dict[int, list[SpreadsheetCell]] = defaultdict(list)
            for cell in all_cells:
                _limit("cell_row", cell.row, XLSX_MAX_LOGICAL_ROW)
                _limit("cell_column", cell.column, XLSX_MAX_LOGICAL_COLUMN)
                rows[cell.row].append(cell)
            for row_cells in rows.values():
                _limit("cells_per_row", len(row_cells), XLSX_MAX_CELLS_PER_ROW)

            raw_merges = list(sheet.merged_cells.ranges)
            _limit("merges", len(raw_merges), XLSX_MAX_MERGES)
            merged = []
            for raw_range in raw_merges:
                region = _range_model(str(raw_range))
                _limit("merge_end_row", region.end_row, XLSX_MAX_LOGICAL_ROW)
                _limit("merge_end_column", region.end_column, XLSX_MAX_LOGICAL_COLUMN)
                range_area(region, "merge")
                merged.append(region)
            merged.sort(
                key=lambda region: (
                    region.start_row,
                    region.start_column,
                    region.end_row,
                    region.end_column,
                    region.a1_range,
                )
            )

            raw_tables = sorted(sheet.tables.values(), key=lambda item: item.name)
            _limit("tables", len(raw_tables), XLSX_MAX_TABLES)
            _limit(
                "table_cell_scans",
                len(raw_tables) * len(all_cells),
                XLSX_MAX_TABLE_CELL_SCANS,
            )
            tables: List[tuple[SpreadsheetTableIdentity, SpreadsheetRange]] = []
            for raw_table in raw_tables:
                region = _range_model(raw_table.ref)
                _limit("table_end_row", region.end_row, XLSX_MAX_LOGICAL_ROW)
                _limit("table_end_column", region.end_column, XLSX_MAX_LOGICAL_COLUMN)
                range_area(region, "table")
                tables.append(
                    (
                        SpreadsheetTableIdentity(
                            name=raw_table.name,
                            display_name=raw_table.displayName,
                            a1_range=raw_table.ref,
                        ),
                        region,
                    )
                )

            sheet_region = SpreadsheetRange(
                a1_range=f"A1:{get_column_letter(max_column)}{max_row}",
                start_row=1,
                start_column=1,
                end_row=max_row,
                end_column=max_column,
            )
            emit(
                StructuredSourceUnit(
                    key=f"sheet:{sheet_ordinal}",
                    ordinal=len(units),
                    kind=StructuredSourceUnitKind.SECTION,
                    text=sheet.title,
                    locator=SpreadsheetLocator(
                        sheet_ordinal=sheet_ordinal,
                        sheet_name=sheet.title,
                        region=sheet_region,
                        cells=all_cells,
                        merged_ranges=merged,
                        defined_tables=[identity for identity, _ in tables],
                    ),
                )
            )

            for row_number in sorted(rows):
                row_cells = rows[row_number]
                min_column = row_cells[0].column
                max_row_column = row_cells[-1].column
                row_region = SpreadsheetRange(
                    a1_range=(
                        f"{get_column_letter(min_column)}{row_number}:"
                        f"{get_column_letter(max_row_column)}{row_number}"
                    ),
                    start_row=row_number,
                    start_column=min_column,
                    end_row=row_number,
                    end_column=max_row_column,
                )
                emit(
                    StructuredSourceUnit(
                        key=f"sheet:{sheet_ordinal}:row:{row_number}",
                        ordinal=len(units),
                        kind=StructuredSourceUnitKind.TABLE_ROW,
                        text=" | ".join(cell.text for cell in row_cells),
                        locator=SpreadsheetLocator(
                            sheet_ordinal=sheet_ordinal,
                            sheet_name=sheet.title,
                            region=row_region,
                            cells=row_cells,
                            merged_ranges=[
                                region for region in merged if contains(row_region, region)
                            ],
                            defined_tables=[],
                        ),
                    )
                )

            for table, table_region in tables:
                table_cells = [
                    cell
                    for cell in all_cells
                    if table_region.start_row <= cell.row <= table_region.end_row
                    and table_region.start_column
                    <= cell.column
                    <= table_region.end_column
                ]
                emit(
                    StructuredSourceUnit(
                        key=f"sheet:{sheet_ordinal}:table:{table.name}",
                        ordinal=len(units),
                        kind=StructuredSourceUnitKind.TABLE_REGION,
                        text="\n".join(cell.text for cell in table_cells),
                        locator=SpreadsheetLocator(
                            sheet_ordinal=sheet_ordinal,
                            sheet_name=sheet.title,
                            region=table_region,
                            cells=table_cells,
                            merged_ranges=[
                                region
                                for region in merged
                                if contains(table_region, region)
                            ],
                            defined_tables=[table],
                        ),
                    )
                )
    finally:
        workbook.close()
    return units


def _as_dataframe(parsed: object) -> pd.DataFrame:
    """ExcelFile.parse is stubbed as DataFrame | dict; we always pass one sheet."""
    if not isinstance(parsed, pd.DataFrame):
        raise TypeError("expected a single worksheet DataFrame")
    return parsed


def _read_sheet_dataframe(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """Read a worksheet into a DataFrame with stable column labels."""
    from openpyxl.utils import get_column_letter

    # XLSX is preprocessed (merge fill); use A/B/C column letters and keep row 1 as data.
    if excel_file.engine == "openpyxl":
        df = _as_dataframe(excel_file.parse(sheet_name=sheet_name, header=None))
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
        return df

    df = _as_dataframe(excel_file.parse(sheet_name=sheet_name, header=0))
    if df.empty:
        df = _as_dataframe(excel_file.parse(sheet_name=sheet_name, header=None))
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    elif any(str(col).startswith("Unnamed:") for col in df.columns):
        df = _as_dataframe(excel_file.parse(sheet_name=sheet_name, header=None))
        df.columns = [get_column_letter(idx + 1) for idx in range(len(df.columns))]
    return df


def _prepare_xlsx_bytes(data: bytes) -> bytes:
    repaired = repair_xlsx_bytes(data)
    if repaired is not None:
        data = repaired
    return fill_merged_cells_xlsx(data)


def _open_excel_file(content: bytes, file_type: str | None = None) -> pd.ExcelFile:
    """Open an Excel workbook with explicit engine selection and fallbacks."""
    data = content
    converted_via_soffice = False

    while True:
        ext = detect_excel_format(data)
        if ext is None:
            if converted_via_soffice:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                )
            try:
                data = normalize_excel_bytes(data, file_type=file_type)
            except ValueError as exc:
                raise ValueError(
                    "Excel file format cannot be determined, you must specify an engine manually."
                ) from exc
            converted_via_soffice = True
            continue

        if ext == "ods":
            converted = convert_excel_to_xlsx_bytes(data, suffix=".ods")
            if converted:
                data = converted
                continue

        engine = engine_for_format(ext)
        if ext == "xlsx":
            data = _prepare_xlsx_bytes(data)
            engine = "openpyxl"
        try:
            return pd.ExcelFile(BytesIO(data), engine=engine)
        except ImportError as exc:
            raise ValueError(
                f"Excel engine {engine!r} is not available for .{ext} files"
            ) from exc
        except KeyError as exc:
            if "sharedStrings.xml" not in str(exc) or engine != "openpyxl":
                raise
            repaired = repair_xlsx_bytes(data)
            if repaired is None:
                raise
            logger.info("Repaired XLSX sharedStrings packaging before parse")
            data = _prepare_xlsx_bytes(repaired)
            continue
        except ValueError as exc:
            if converted_via_soffice or "cannot be determined" not in str(exc):
                raise
            try:
                data = normalize_excel_bytes(content, file_type=file_type)
            except ValueError:
                raise
            converted_via_soffice = True
            continue


if __name__ == "__main__":
    # Example usage: Parse an Excel file and display results
    logging.basicConfig(level=logging.DEBUG)

    # Specify the path to your Excel file
    your_file = "/path/to/your/file.xlsx"
    parser = ExcelParser()

    # Read and parse the Excel file
    with open(your_file, "rb") as f:
        content = f.read()
        document = parser.parse_into_text(content)

        # Display the full document content
        logger.error(document.content)

        # Display the first chunk as an example
        for chunk in document.chunks:
            logger.error(chunk.content)
            break  # Only show the first chunk
